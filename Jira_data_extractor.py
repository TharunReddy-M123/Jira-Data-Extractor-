import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import pytz
import re
import os  # For file operations

def parse_jira_date(date_str):
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f%z")
        return dt.astimezone(pytz.utc).replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return ""

def format_jira_time(seconds):
    if pd.isnull(seconds) or seconds is None or seconds <= 0:
        return None
    if seconds <= 28800:  # Up to 8 hours
        hours = seconds // 3600
        remaining = seconds % 3600
        minutes = remaining // 60
        secs = remaining % 60
        return f"{int(hours):02}:{int(minutes):02}:{int(secs):02}"
    else:  # More than 8 hours
        days = seconds // 28800
        remaining_seconds = seconds % 28800
        hours = remaining_seconds // 3600
        remaining_seconds %= 3600
        minutes = remaining_seconds // 60
        secs = remaining_seconds % 60
        days_str = "day" if days == 1 else "days"
        return f"{int(days)} {days_str}, {int(hours):02}:{int(minutes):02}:{int(secs):02}"

def format_timeoriginalestimate(seconds):
    if pd.isnull(seconds) or seconds is None or seconds <= 0:
        return "None"
    workday_seconds = 28800  # 8 hours in seconds
    days = seconds // workday_seconds
    remaining_seconds = seconds % workday_seconds
    
    hours = remaining_seconds // 3600
    remaining_seconds %= 3600
    minutes = remaining_seconds // 60
    secs = remaining_seconds % 60
    
    days_str = "day" if days == 1 else "days"
    
    if days > 0:
        return f"{int(days)} {days_str}, {int(hours):02}:{int(minutes):02}:{int(secs):02}"
    else:
        return f"{int(hours):02}:{int(minutes):02}:{int(secs):02}"

def sprint(data):
    try:
        updated_list = []
        for fields in data:
            sprint_data = str(fields.get("customfield_10018", ""))
            match = re.search(r"name=(.*?),", sprint_data)
            updated_list.append(match.group(1) if match else "None")
        return updated_list
    except Exception as e:
        print(f"\nSprint Error: {e}")
        return ["None"] * len(data)

def bug_source(data):
    try:
        updated_list = []
        for fields in data:
            cf = fields.get("customfield_11504")
            if cf and 'value' in cf:
                updated_list.append(cf['value'])
            else:
                updated_list.append("None")
        return updated_list
    except Exception as e:
        print(f"\nBug Source Error: {e}")
        return ["None"] * len(data)

def timespent(data):
    try:
        updated_list = []
        for name in data:
            if name.get('timespent') is not None:
                updated_list.append(format_jira_time(name['timespent']))
            else:
                updated_list.append("None")
        return updated_list
    except Exception as e:
        print(f"\nError(timespent) --> {e}")
        return ["None"] * len(data)

# Jira API config
url = "https://jira.rampgroup.com/rest/api/2/search?jql=project=GLOW&maxResults=500"
auth = ("tharun.morreddygari@rampgroup.com", "Mkumar#12345")
headers = {"Accept": "application/json"}

response = requests.get(url, auth=auth, headers=headers)

if response.status_code == 200:
    data = response.json()
    issues = data.get("issues", [])
    fields_list = [issue.get("fields", {}) for issue in issues]
    
    sprint_names = sprint(fields_list)
    bug_sources = bug_source(fields_list)
    timespent_data = timespent(fields_list)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jira Issues"
    
    # Original headers list
    headers = [
        'index', 'Assignee_name', 'Project_name', 'Issue_type', 'Issue_key',
        'Priority', 'Status', 'Reporter_name', 'Creator_name', 'Resolution',
        'Resolution_date', 'Activity_date', 'Created_date', 'Updated_date',
        'Link', 'timeoriginalestimate', 'timeestimate',
        'aggregatetimeoriginalestimate', 'timespent', 'aggregatetimespent',
        'Sprint', 'bug_source', 'labels', 'Time_Spent', 'Estimated_Time'
    ]

    # Columns to remove in Excel output
    columns_to_remove = ['Resolution_date', 'aggregatetimeoriginalestimate', 'aggregatetimespent']
    excel_headers = [h for h in headers if h not in columns_to_remove]

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Write headers with bold font
    for col_num, header in enumerate(excel_headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = bold_font

    issue_data = []
    
    for idx, (issue, sprint_name, bug_src, timespent_value) in enumerate(zip(issues, sprint_names, bug_sources, timespent_data), start=1):
        fields = issue.get("fields", {})
        
        created_date = parse_jira_date(fields.get("created"))
        updated_date = parse_jira_date(fields.get("updated"))
        resolution_date = parse_jira_date(fields.get("resolutiondate"))

        time_estimate = fields.get("timeoriginalestimate")
        formatted_time_estimate = format_timeoriginalestimate(time_estimate)

        issue_data.append([
            idx,
            (fields.get("assignee") or {}).get("emailAddress", "None"),
            (fields.get("project") or {}).get("name", ""),
            (fields.get("issuetype") or {}).get("name", ""),
            issue.get("key", ""),
            (fields.get("priority") or {}).get("name", ""),
            (fields.get("status") or {}).get("name", ""),
            (fields.get("reporter") or {}).get("displayName", ""),
            (fields.get("creator") or {}).get("displayName", ""),
            (fields.get("resolution") or {}).get("name", "None"),
            resolution_date,
            updated_date,
            created_date,
            updated_date,
            f"{url.split('/rest')[0]}/browse/{issue.get('key', '')}",
            formatted_time_estimate,
            fields.get("timeestimate", 0),
            fields.get("aggregatetimeoriginalestimate", 0),
            timespent_value,
            fields.get("aggregatetimespent", 0),
            sprint_name,
            bug_src,
            ", ".join(fields.get("labels", [])),
            None,
            None
        ])

    df = pd.DataFrame(issue_data, columns=headers)

    # Drop unwanted columns before exporting to Excel
    df_excel = df.drop(columns=columns_to_remove)

    # Convert date columns to datetime
    df_excel['Activity_date'] = pd.to_datetime(df_excel['Activity_date'], errors='coerce')
    df_excel['Created_date'] = pd.to_datetime(df_excel['Created_date'], errors='coerce')
    df_excel['Updated_date'] = pd.to_datetime(df_excel['Updated_date'], errors='coerce')

    max_days = 5

    df_excel['Time_Spent'] = df_excel.apply(
        lambda row: row['Activity_date'] - row['Created_date']
        if pd.notnull(row['Activity_date']) else row['Updated_date'] - row['Created_date'], 
        axis=1
    )

    df_excel['Time_Spent'] = df_excel['Time_Spent'].apply(
        lambda x: x - timedelta(days=((x.days // 7) * 2 + (x.days % 7 >= 5))) 
        if pd.notnull(x) else x
    )

    df_excel['Estimated_Time'] = pd.to_timedelta(df_excel['Time_Spent']) - pd.Timedelta(days=max_days)
    df_excel.loc[(df_excel['Status'] == 'Done') & (df_excel['Estimated_Time'] < pd.Timedelta(0)), 'Estimated_Time'] = 'In Time'

    df_excel['Time_Spent'] = df_excel['Time_Spent'].apply(lambda x: f"{x.days} days" if pd.notnull(x) else x)
    df_excel['Estimated_Time'] = df_excel['Estimated_Time'].apply(
        lambda x: x if x == 'In Time' else f"{x.days} days" if pd.notnull(x) else x
    )

    output_path = r"D:\New_Jira_Data_Extractor.xlsx"
    if os.path.exists(output_path):
        os.remove(output_path)

    # Write DataFrame rows to Excel, starting from row 2 (row 1 is header)
    for r_idx, row in enumerate(dataframe_to_rows(df_excel, header=False, index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            # Bold and center-align the 'index' column values (first column)
            if c_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(output_path)
    print(f"Data exported successfully to {output_path}!")
else:
    print(f"Error {response.status_code}: {response.text}")
